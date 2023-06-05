"""
[for .xlsx files with LSB choice, sometime it works sometimes it doesnt :')]
"""
import openpyxl
import os


def hide_data_in_excel(base_file, payload, num_lsb):
    wb = openpyxl.load_workbook(base_file)
    sheet = wb.active

    binary_data = ''.join(format(ord(char), '08b') for char in payload)
    binary_data += '0' * num_lsb  # Pad with zeros to match LSB size

    index = 0
    for row in sheet.iter_rows():
        for cell in row:
            if index >= len(binary_data):
                break
            original_value = cell.value
            if original_value is None:
                original_value = ''
            elif isinstance(original_value, float):
                original_value = str(original_value)
            elif isinstance(original_value, str):
                modified_value = ''
                for char in original_value:
                    binary_value = format(ord(char), '08b')
                    modified_value += binary_value[:-num_lsb] + binary_data[index:index+num_lsb]
                    index += num_lsb
                cell.value = ''.join(chr(int(modified_value[i:i+8], 2)) for i in range(0, len(modified_value), 8))

    #root, ext = os.path.splitext(base_file)
    #dir_path = os.path.dirname(root)
    #base_name = os.path.splitext(os.path.basename(base_file))[0]
    #output = base_name + f"_encoded" + ext
    #encoded_file = os.path.join(dir_path, output)
    encoded_file = 'encoded_excel.xlsx'
    wb.save(encoded_file)
    print('written to:', encoded_file)
    return encoded_file



def extract_data_from_excel(file_path, num_lsb):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    extracted_data = ''
    for row in sheet.iter_rows():
        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                continue
            elif isinstance(cell_value, float):
                cell_value = str(cell_value)
            elif isinstance(cell_value, str):
                for char in cell_value:
                    binary_value = format(ord(char), '08b')
                    extracted_data += binary_value[-num_lsb:]

    extracted_message = ''
    for i in range(0, len(extracted_data), 8):
        byte = extracted_data[i:i+8]
        if byte:
            extracted_message += chr(int(byte, 2))

    print("decoded thingy:", extracted_message)
    return extracted_message



if __name__ == '__main__':
    base_file = r"C:\Users\USER\PycharmProjects\stegosaurus\Excel\untitled.xlsx"
    stego_file = r"C:\Users\USER\PycharmProjects\stegosaurus\Excel\untitled_encoded.xlsx"
    payload = "jed was definitely here"
    num_lsb = 4  # (e.g., 0, 1, 2, ..., 5)

    hide_data_in_excel(base_file, payload, num_lsb)
    extract_data_from_excel(stego_file, num_lsb)
